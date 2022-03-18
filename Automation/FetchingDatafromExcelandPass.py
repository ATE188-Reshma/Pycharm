import time
import openpyxl
import pytest
import pyscreenshot
from selenium import webdriver

def xldata():
    excel_path = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample1.xlsx"
    open_workbook = openpyxl.load_workbook(excel_path)
    open_worksheet = open_workbook["Test Case"]
    max_row = open_worksheet.max_row
    max_column = open_worksheet.max_column
    kkk = []
    print(max_row, max_column)
    # Bcoz range starts from 0. So, Max row +1.
    for i in range(2, max_row + 1):
        kk = []
        for j in range(1, max_column + 1):
            data = open_worksheet.cell(i, j).value
            kk.insert(j, data)
        print(kk)
        kkk.insert(i, kk)
    return kkk


@pytest.mark.parametrize("var1,var2", xldata())
def test_eval1(var1, var2):
    driver = webdriver.Chrome(executable_path="C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\Automation\\chromedriver.exe")
    driver.get("http://192.168.0.199:9091/QuaLISWeb/#/login")
    time.sleep(4)
    print(var2)
    driver.find_element_by_xpath("//*[@name='sloginid']").send_keys(var1)
    time.sleep(4)
    driver.find_element_by_xpath("//*[@name='spassword']").send_keys(var2)
    time.sleep(4)
    pyscreenshot.grab().save("D:\\Resh JPDC\\Selenium_Python\\a.png")
    driver.find_element_by_xpath("//*[text()='Login']").click()

