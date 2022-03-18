import pypyodbc
import openpyxl
dbcredentials="Driver={SQL Server};Server=AGD55\SQLSERVER2017;Database=JPDC;UID=sa;PWD=admin@123"
dbconnect=pypyodbc.connect(dbcredentials)
print(dbconnect)
cursor=dbconnect.cursor()
selectquery="select * from passwordpolicy"
update="SELECT name FROM sys.columns WHERE object_id = OBJECT_ID('passwordpolicy')"
select=cursor.execute(update)
print(select)

initialRow = 1
#secondrow=2

for x in select:
    initialColumn = 1

    for y in x:


        location = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\DBfetching1.xlsx"

        workbook = openpyxl.load_workbook(location)
        sheet = workbook["comparison"]
        print(y)

        df = sheet.cell(initialColumn, initialRow).value = y
        workbook.save(location)
        initialColumn = initialColumn + 1
    initialRow = initialRow + 1

select=cursor.execute(selectquery)
initialRow = 2
#secondrow=2

for x in select:
    initialColumn = 1

    for y in x:

        workbook = openpyxl.load_workbook(location)
        sheet = workbook["comparison"]
        print(y)

        df = sheet.cell(initialRow, initialColumn).value = y
        workbook.save(location)
        initialColumn = initialColumn + 1
    initialRow = initialRow + 1


