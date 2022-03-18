import time
import openpyxl
import pytest
from selenium import webdriver

class Testworkbook():
    @pytest.fixture(scope="session")
    def test_parent(self):
        excel_path = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx"
        global open_workbook, open_worksheet
        open_workbook = openpyxl.load_workbook(excel_path)
        open_worksheet = open_workbook["comparison"]
        maxrow = open_worksheet.max_row
        print(maxrow)
        global a,b,c,d
        a = open_worksheet.cell(3,2).value
        b = open_worksheet.cell(3,3).value
        print(a)
        print(b)
        c = open_worksheet.cell(5,4).value
        d = open_worksheet.cell(5,5).value
        print(c)
        print(d)

    def test_child(self, test_parent):
        driver = webdriver.Chrome(executable_path="C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\Automation\\chromedriver.exe")
        driver.maximize_window()
        driver.implicitly_wait(15)
        driver.get("http://192.168.0.199:9091/QuaLISWeb/#/login")
        loginid = driver.find_element_by_xpath("//*[@name='sloginid']")
        loginid.click()
        time.sleep(3)
        loginid.send_keys("cdolman")
        time.sleep(4)
        passwd = driver.find_element_by_xpath("//*[@name='spassword']")
        time.sleep(3)
        passwd.click()
        passwd.send_keys("123")
        time.sleep(5)

        role = driver.find_element_by_xpath("//div[text()='Admin']").text
        print("\n")
        print(role)
        logintype = driver.find_element_by_xpath("//div[text()='Internal']").text
        print(logintype)

        driver.find_element_by_xpath("//*[text()='Login']").click()
        time.sleep(4)

        if (role==a and logintype==b):
            open_worksheet.cell(3, 7).value = "Pass"
            open_workbook.save("D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx")
            open_workbook.close()
            print("PASS\n")

        transaction = driver.find_element_by_xpath("//div[text()='Transaction']").text
        print(transaction)
        samplereceiving = driver.find_element_by_xpath("//span[text()='Sample Receiving']").text
        print(samplereceiving)

        if (transaction==c and samplereceiving==d):
            open_worksheet.cell(5, 7).value = "Pass"
            open_workbook.save("D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx")
            open_workbook.close()
            print("PASS")
