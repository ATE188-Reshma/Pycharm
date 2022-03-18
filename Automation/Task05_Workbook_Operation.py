import time
import openpyxl
from selenium import webdriver

excel_path = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx"
open_workbook = openpyxl.load_workbook(excel_path)
open_worksheet = open_workbook["comparison"]
maxrow = open_worksheet.max_row
print(maxrow)

a = open_worksheet.cell(3,2).value
b = open_worksheet.cell(3,3).value
print(a)
print(b)
c = open_worksheet.cell(5,4).value
d = open_worksheet.cell(5,5).value
print(c)
print(d)

driver = webdriver.Chrome(executable_path="../Automation/chromedriver.exe")

driver.maximize_window()
driver.implicitly_wait(15)
driver.get("http://192.168.0.199:9091/QuaLISWeb/#/login")

driver.find_element_by_xpath("//*[@name='sloginid']").send_keys("cdolman")
time.sleep(4)
driver.find_element_by_xpath("//*[@name='spassword']").send_keys("123")
time.sleep(5)

role = driver.find_element_by_xpath("//div[text()='Admin']").text
print(role)
logintype = driver.find_element_by_xpath("//div[text()='Internal']").text
print(logintype)

driver.find_element_by_xpath("//*[text()='Login']").click()
time.sleep(4)

if (role==a and logintype==b):
    open_worksheet.cell(3, 7).value = "Pass"
    open_workbook.save("D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx")
    open_workbook.close()
    print("PASS")

transaction = driver.find_element_by_xpath("//div[text()='Transaction']").text
print(transaction)
samplereceiving = driver.find_element_by_xpath("//span[text()='Sample Receiving']").text
print(samplereceiving)

if (transaction==c and samplereceiving==d):
    open_worksheet.cell(5, 7).value = "Pass"
    open_workbook.save("D:\\Resh JPDC\\Selenium_Python\\K_Notes\\sample.xlsx")
    open_workbook.close()
    print("\nPASS")








