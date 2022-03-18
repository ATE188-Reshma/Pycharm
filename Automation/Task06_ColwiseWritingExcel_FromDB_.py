import pypyodbc
import openpyxl

dbcredentials = "Driver={SQL Server};Server=AGD55\SQLSERVER2017;Database=IRSHALIMSPROD_31Dec;UID=sa;PWD=admin@123"
dbconnect = pypyodbc.connect(dbcredentials)
cursor = dbconnect.cursor()

selectquery = "select nusercode from users"
select = cursor.execute(selectquery)
print(select)

path = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\DBfetching1.xlsx"
workbook = openpyxl.load_workbook(path)
worksheet = workbook["comparison"]
row=1
column=1
for i in select:
    print(i)
    n=str(i)
    worksheet.cell(row, column).value = n
    row += 1
    workbook.save(path)
    workbook.close()

selectquery = "select sloginid from users"
select = cursor.execute(selectquery)
print(select)

path = "D:\\Resh JPDC\\Selenium_Python\\K_Notes\\DBfetching1.xlsx"
workbook = openpyxl.load_workbook(path)
worksheet = workbook["comparison"]
row=1
column=2
for i in select:
    print(i)
    n=str(i)
    worksheet.cell(row, column).value = n
    row += 1
    workbook.save(path)
    workbook.close()
















