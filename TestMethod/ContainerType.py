import time
from configparser import ConfigParser
from os import listdir

import openpyxl
from loguru import logger
from selenium.webdriver.common.by import By

import JDBC
from TestMethod import AuditTrial
from TestMethod.AuditTrial import auditTrail
from Utiliser import ReusableMethod

objectRepository = ConfigParser()
objectRepository.read("C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\ObjectRepository\\Element_ContainerType.ini")

objectRepository2 = ConfigParser()
objectRepository2.read("C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\ObjectRepository\\Element_PageCount.ini")

InputData = ConfigParser()
InputData.read("C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\InputData\\Data_ContainerType.ini")

objectRepository1=ConfigParser()
objectRepository1.read("C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\ObjectRepository\\Element_LaunchLims_BasicAction.ini")

objectRepository3=ConfigParser()
objectRepository3.read("C:\\Users\\ate142\\PycharmProjects\\pythonProject1\\InputData\\Data_LaunchLims.ini")

global id

id=objectRepository3.get("credentials Values", "inputid")

def containerType_Prequesite(driver, basic, module_key):
    masterxpath = objectRepository1.get(basic, "master")
    basemasterxpath = objectRepository.get(module_key, "basemaster")
    containertypexpath = objectRepository.get(module_key, "containertype")

    try:
        time.sleep(8)
        ReusableMethod.clickXpath(driver,masterxpath)
        logger.info("Master module clicked")

        try:
            ReusableMethod.clickXpath(driver, basemasterxpath)
            logger.info("Basemaster module clicked")

            try:
                ReusableMethod.clickXpath(driver, containertypexpath)
                logger.info("Containertype screen clicked")

            except Exception as e:
                logger.error("Containertype screen has not clicked, Exception occurred" + str(e))

        except Exception as e:
            logger.error("Basemaster module has not clicked, Exception occurred" + str(e))

    except Exception as e:
        logger.error("Master module has not clicked, Exception occurred" + str(e))


def containerType_Add(driver, loc_key, input_key):
    addxpath = objectRepository.get(loc_key, "add")
    containertypexpath = objectRepository.get(loc_key, "containertype")
    descriptionxpath = objectRepository.get(loc_key, "description")
    savexpath = objectRepository.get(loc_key, "save")
    containertypevalue = InputData.get(input_key, "inputcontainertype")
    descriptionvalue = InputData.get(input_key, "inputdescription")

    try:
        driver.implicitly_wait(10)
        time.sleep(3)
        ReusableMethod.clickXpath(driver, addxpath)
        logger.info("Add button clicked")

        try:
            ReusableMethod.sendKeysXpath(driver, containertypexpath, containertypevalue)
            logger.info("Entered ContainerType")

            try:
                ReusableMethod.sendKeysXpath(driver, descriptionxpath, descriptionvalue)
                logger.info("Entered ContainerTypeDescription")

                try:
                    ReusableMethod.clickXpath(driver, savexpath)
                    logger.info("ContainerType Added")

                except Exception as e:
                    logger.error("ContainerType has not Added, Exception occurred" + str(e))

            except Exception as e:
                logger.error("ContainerTypeDescription value has not passed, Exception occurred" + str(e))

        except Exception as e:
            logger.error("ContainerType value has not passed, Exception occurred" + str(e))

    except Exception as e:
        logger.error("Add button has not clicked, Exception occurred" + str(e))


def containerType_Edit(driver, loc_key, input_key, tag):
    containertypexpath = objectRepository.get(loc_key, "containertype")
    descriptionxpath = objectRepository.get(loc_key, "description")
    savexpath = objectRepository.get(loc_key, "save")
    containertype_editvalue = InputData.get(input_key, "editinputcontainertype")
    description_editvalue = InputData.get(input_key, "editinputdescription")

    time.sleep(7)

    try:
        t = driver.find_elements(By.TAG_NAME, tag)
        tt = len(t)

        # input("enter containertype with description: ")
        # pass the record to edit
        container = "Soap Container container with 2 compartments"
        for i in range(1, tt):
            ttt = t[i].text
            print(ttt)
            if container == ttt:
                print("containertype count is", i, ttt)
                print(i)
                l = str(i)
                edit = "(//span[@data-tip='Edit'])[" + l + "]"
                print(edit)
                driver.find_element(By.XPATH, edit).click()
                logger.info("Edit button clicked")
                break


        try:
            time.sleep(3)
            driver.find_element(By.XPATH, containertypexpath).clear()
            time.sleep(2)
            ReusableMethod.sendKeysXpath(driver, containertypexpath, containertype_editvalue)
            driver.find_element(By.XPATH, descriptionxpath).clear()
            time.sleep(2)
            ReusableMethod.sendKeysXpath(driver, descriptionxpath, description_editvalue)
            logger.info("ContainerType value passed for Edit")

            try:
                ReusableMethod.clickXpath(driver, savexpath)
                logger.info("ContainerType Edited")

            except Exception as e:
                logger.error("ContainerType has not Edited, Exception occurred" + str(e))


        except Exception as e:
            logger.error("ContainerType value has not passed for Edit, Exception occurred" + str(e))


    except Exception as e:
        logger.error("Edit button has not clicked, Exception occured" + str(e))


def containerType_Delete(driver, loc_key, tag):
    containertypedeleteokxpath = objectRepository.get(loc_key, "deletepopupokbutton")
    containertypedeletecancelxpath = objectRepository.get(loc_key, "deletepopupcancelbutton")

    time.sleep(7)

    try:
        q = driver.find_elements(By.TAG_NAME, tag)
        qq = len(q)

        # input("enter containertype with description: ")
        container = "Mat matrix container"
        for i in range(1, qq):
            qqq = q[i].text
            print(qqq)
            if container == qqq:
                print("containertype count is", i, qqq)
                print(i)
                m = str(i)
                delete = "(//span[@data-tip='Delete'])[" + m + "]"
                print(delete)
                driver.find_element(By.XPATH, delete).click()
                break
        logger.info("Delete button clicked")

        try:
            time.sleep(10)
            ReusableMethod.clickXpath(driver, containertypedeleteokxpath)
            logger.info("ContainerType Deleted")

        except Exception as e:
            logger.error("ContainerType has not Deleted, Exception occurred" + str(e))


    except Exception as e:
        logger.error("Delete button has not clicked, Exception occurred" + str(e))



def containerType_search(driver, loc_key, filterdata):
    containerType_filtericon = objectRepository.get(loc_key, "containertypefiltericon")
    containerType_filtericon_contains = objectRepository.get(loc_key, "containertypefiltericon_contains")
    containerType_filtericon_containstextfield = objectRepository.get(loc_key, "containertypefiltericon_contains_textfield")
    containerType_filtericon_and = objectRepository.get(loc_key, "containertypefiltericon_And")
    containerType_filtericon_and_contains = objectRepository.get(loc_key, "containertypefiltericon_Andcontains")
    containerType_filtericon_and_contains_textfield = objectRepository.get(loc_key, "containertypefiltericon_Andcontains_textfield")
    containerType_filtericon_filterbutton = objectRepository.get(loc_key, "containertypefiltericon_filterbutton")
    containerType_filtericon_clearbutton = objectRepository.get(loc_key, "containertypefiltericon_clearbutton")
    containerType_filtericon_containstextfieldvalue = InputData.get(filterdata, "containsinput")
    containerType_filtericon_and_contains_textfieldvalue = InputData.get(filterdata, "andcontainsinput")
    containerType_refreshbutton = objectRepository.get(loc_key, "refreshbutton")

    time.sleep(7)

    try:
        ReusableMethod.clickXpath(driver, containerType_filtericon)

        ReusableMethod.clickXpath(driver, containerType_filtericon_contains)
        time.sleep(2)

        element = driver.find_element(By.XPATH, "//*[text()='Is equal to']")
        driver.execute_script("arguments[0].scrollIntoView();", element)
        time.sleep(2)
        element.click()

        time.sleep(2)
        ReusableMethod.sendKeysXpath(driver, containerType_filtericon_containstextfield,
                                     containerType_filtericon_containstextfieldvalue)

        ReusableMethod.clickXpath(driver, containerType_filtericon_and)
        time.sleep(2)

        element = driver.find_element(By.XPATH, "//*[text()='Or']")
        driver.execute_script("arguments[0].scrollIntoView();", element)
        time.sleep(2)
        element.click()

        time.sleep(2)
        ReusableMethod.clickXpath(driver, containerType_filtericon_and_contains)
        time.sleep(2)

        element = driver.find_element(By.XPATH, "//*[text()='Starts with']")
        driver.execute_script("arguments[0].scrollIntoView();", element)
        time.sleep(2)
        element.click()

        time.sleep(2)
        ReusableMethod.sendKeysXpath(driver, containerType_filtericon_and_contains_textfield,
                                     containerType_filtericon_and_contains_textfieldvalue)

        time.sleep(2)
        ReusableMethod.clickXpath(driver, containerType_filtericon_filterbutton)

        logger.info("ContainerType data filtered")

    except Exception as e :
        logger.error("ContainerType data not filtered" +str(e))


    time.sleep(2)

    try:
        ReusableMethod.clickXpath(driver, containerType_refreshbutton)
        logger.info("ContainerType Screen refreshed")

    except Exception as e:
        logger.error("ContainerType Screen has not refreshed")



def containerType_downloadPDF(driver, loc_key):
    downloadpdf = objectRepository.get(loc_key, "downloadPDF")

    time.sleep(3)
    ReusableMethod.clickXpath(driver, downloadpdf)
    logger.info("ContainerType PDF downloaded")


    # filepath = "C:\\Users\\ate142\\Downloads"
    # # openfile = open(filepath)
    # a = [i for i in listdir(filepath)]
    # for j in range(len(a)):
    #     b = a[j]
    #     print(filepath+"\\"+b)

def containerType_downloadExcel(driver, loc_key):
    downloadexcel = objectRepository.get(loc_key, "downloadExcel")

    time.sleep(5)

    downloadPath="C:\\Users\\ate142\\Downloads\\"

    oldFileList = [f for f in listdir(downloadPath)]

    beforeCount=len(oldFileList)

    ReusableMethod.clickXpath(driver, downloadexcel)

    time.sleep(5)

    newFileList=[f for f in listdir(downloadPath)]

    afterCount=len(newFileList)

    oldFilees=""

    newFile=None

    if afterCount==beforeCount+1:
        logger.info("The file is downloaded")

        for i in oldFileList:
            oldFilees=oldFilees+i

        for i in newFileList:
            if oldFilees.__contains__(i):
                pass
            else:
                newFile=i

        excelFileLocation=downloadPath+newFile

        open_workbook = openpyxl.load_workbook(excelFileLocation)

        open_worksheet = open_workbook["Sheet1"]

        nameHeader = open_worksheet.cell(row=1, column=1).value

        descriptionHeader = open_worksheet.cell(row=1, column=2).value

        if nameHeader == "Container Type":
            logger.info("The Container Type name header is displayed properly")
        else:
            logger.error("The Container Type name header is not displayed properly")

        if descriptionHeader == "Description":
            logger.info("The Container Type description header is displayed properly")
        else:
            logger.error("The Container Type name header is not displayed properly")

        exceldataCount = open_worksheet.max_row - 1

        dbCount=JDBC.containerTypeCount()

        if   exceldataCount==dbCount:
            logger.info("Number of data is matched")
        else:
            logger.error("Number of data is mis-matched")

    else:
        logger.error("The  file is not downloaded")




    time.sleep(3)

    # to get the page count
def count(driver, loc_key, xpath):
        # change xpath as required
        containertypepagecountxpath = objectRepository2.get(loc_key, xpath)

        containertypepagecount = driver.find_element(By.XPATH, containertypepagecountxpath).text
        print(containertypepagecount)

        individualtext = containertypepagecount.split(' ')
        print(individualtext)

        count = individualtext[4]

        print(count)

        count = int(count)

        return count

    # Page count validation
def countvalidation(screen, beforecount, aftercount):
        if aftercount == beforecount + 1:
            logger.info(screen + " count increased by 1")

        elif aftercount > beforecount + 1:
            logger.info(screen + " count increased more than 1")

        elif aftercount == beforecount:
            logger.info(screen + " count has not increased")

        elif aftercount == beforecount - 1:
            logger.info(screen + " count decreased by 1")

        elif aftercount < beforecount - 1:
            logger.info(screen + " count decreased more than 1")

def indexValidateAdd(driver, input_key):
        # containertypeindexxpath = objectRepository1.get(loc_key, "containertypeindex")
        containertypevalue = InputData.get(input_key, "inputcontainertype")
        descriptionvalue = InputData.get(input_key, "inputdescription")
        # refresh = objectRepository1.get(loc_key, "refreshbutton")

        containertypeindex = driver.find_elements(By.TAG_NAME, "tr")
        tt = len(containertypeindex)

        containerdata = containertypevalue + " " + descriptionvalue

        # ReusableMethod.clickXpath(driver, refresh)
        # time.sleep(4)

        if containertypeindex[1].text == containerdata:
            logger.info("Added Container Type displayed in first index")

        else:
            # ReusableMethod.clickXpath(driver,refresh)
            # time.sleep(2)
            for i in range(1, tt):
                ttt = containertypeindex[i].text
                print(ttt)

                if containerdata == ttt:
                    print("Container Type displayed in index" + str(i))


    # Add
def containerTypeCountIndexAddValidation(driver):
        beforecount = count(driver, "containertype", "containertypepagecount")
        containerType_Add(driver, "screen locator", "screen value")
        time.sleep(5)
        aftercount = count(driver, "containertype", "containertypepagecount")
        countvalidation("containerType", beforecount, aftercount)
        time.sleep(5)
        indexValidateAdd(driver, "screen value")
        time.sleep(5)


    #Edit
def containerTypeCountIndexEditValidation(driver):
        beforecount = count(driver, "containertype", "containertypepagecount")
        containerType_Edit(driver, "screen locator", "screen value", "tr")
        time.sleep(5)
        aftercount = count(driver, "containertype", "containertypepagecount")
        countvalidation("containerType", beforecount, aftercount)
        time.sleep(5)



    #Delete
def containerTypeCountIndexDeleteValidation(driver):
        beforecount = count(driver, "containertype", "containertypepagecount")
        containerType_Delete(driver, "screen locator", "tr")
        time.sleep(5)
        aftercount = count(driver, "containertype", "containertypepagecount")
        countvalidation("containerType", beforecount, aftercount)
        time.sleep(5)

    # Add
def containerTypeAddAuditTrail(driver):
        beforeCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(2)
        containerType_Prequesite(driver, "basic", "module screen")

        time.sleep(2)
        containerType_Add(driver, "screen locator", "screen value")

        time.sleep(3)
        afterCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(3)
        auditTrail(driver, afterCount, beforeCount, "ADD CONTAINER TYPE", "Carl Dolman", "Admin",
                   "Container Type: Soap Container;Description: container with 2 compartments;", "SYSTEM")

    # Edit
def containerTypeEditAuditTrail(driver):
        time.sleep(4)

        beforeCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(2)
        containerType_Prequesite(driver, "basic", "module screen")

        time.sleep(2)
        containerType_Edit(driver, "screen locator", "screen value", "tr")

        time.sleep(2)
        afterCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(3)
        auditTrail(driver, afterCount, beforeCount, "EDIT CONTAINER TYPE", "Carl Dolman", "Admin",
                   "Container Type: Soap Container-> Mat;Description: container with 2 compartments-> matrix container;",
                   "SYSTEM")

    # Delete
def containerTypeDeleteAuditTrail(driver):
        time.sleep(4)

        beforeCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(2)
        containerType_Prequesite(driver, "basic", "module screen")

        time.sleep(2)
        containerType_Delete(driver, "screen locator", "tr")

        time.sleep(2)
        afterCount = AuditTrial.auditTrailCount(driver, "audittrail")

        time.sleep(3)
        auditTrail(driver, afterCount, beforeCount, "DELETE CONTAINER TYPE", "Carl Dolman", "Admin",
                   "Container Type: Mat;Description: matrix container;", "SYSTEM")








